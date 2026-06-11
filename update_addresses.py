#!/usr/bin/env python3
"""Config-driven door-plate (門牌) address updater for Taiwan counties.

Downloads a county's door-plate coordinate dataset from data.gov.tw /
data.nat.gov.tw, converts TWD97/TM2 (EPSG:3826) coordinates to WGS84, splits
road sections, and regenerates the per-road CSV files under roads/.

Each county differs only in: dataset id, source column header names, and the
town-code scheme. Those live in COUNTIES below; the pipeline is shared.

Usage:
    pip install requests pyproj openpyxl   # openpyxl only for XLSX-only counties
    python update_addresses.py 新北市
    python update_addresses.py --list
    python update_addresses.py --all
"""

import argparse
import csv
import io
import re
import ssl
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import requests
from requests.adapters import HTTPAdapter
from urllib3.poolmanager import PoolManager

ROOT = Path(__file__).resolve().parent
ROADS_DIR = ROOT / "roads"
AREA_CSV = ROOT / "area_2014.csv"

CSV_HEADER = [
    "FULL_ADDR", "COUNTY", "TOWN", "VILLAGE", "NEIGHBORHOOD",
    "ROAD", "SECTION", "LANE", "ALLEY", "SUB_ALLEY", "TONG",
    "NUMBER", "X", "Y",
]

# Trailing road section, e.g. 民生東路五段 -> (民生東路, 五段)
_SECTION_RE = re.compile(r"^(.*?)([一二三四五六七八九十\d]+段)$")


# ── County configuration ───────────────────────────────────────────────────────

@dataclass
class CountyConfig:
    name: str            # 新北市
    code: str            # 65  (road-file prefix and COUNTY column)
    dataset_id: int | None  # data.gov.tw id; None when using url/page_url instead
    # canonical field -> source CSV header name
    columns: dict
    domain: str = "data.gov.tw"   # primary; falls back to data.nat.gov.tw
    town_scheme: str = "auto"     # "auto"=detect by code length, "x10", "name"
    epsg: str = "EPSG:3826"       # source coordinate system
    fmt: str = "csv"              # source distribution format: "csv" or "xlsx"
    multi: bool = False           # merge ALL matching distributions (parts), not
                                  # just the first — only when they are genuine
                                  # split parts (e.g. 雲林 XLSX), not redundant copies
    url: str | None = None        # explicit download URL (skip data.gov.tw API)
    page_url: str | None = None   # HTML page to scrape for the latest data link
    page_re: str | None = None    # regex capturing the link href on page_url
    addr_col: str | None = None   # full-address column to parse (e.g. 新竹市 地址)


# Standard Chinese-header layout (most counties)
_ZH = dict(town="鄉鎮市區代碼", village="村里", nei="鄰", street="街路段",
           lane="巷", alley="弄", number="號", x="橫座標", y="縱座標")


def _zh(**kw) -> dict:
    """_ZH template with column-name overrides."""
    return {**_ZH, **kw}


COUNTIES = {
    # ── 六都 ──────────────────────────────────────────────────────────────────
    "臺北市": CountyConfig("臺北市", "63", 155472, _ZH),
    "新北市": CountyConfig(
        "新北市", "65", 168887,
        columns=dict(town="areacode", village="village", nei="neighbor",
                     street="street、road、section", lane="lane", alley="alley",
                     number="number", x="x_3826", y="y_3826"),
        town_scheme="x10",
    ),
    "桃園市": CountyConfig("桃園市", "68", 157689, _ZH, domain="data.nat.gov.tw"),
    # 臺中市 (66, 169806) already maintained by a separate script
    "臺南市": CountyConfig("臺南市", "67", 120044, _zh(street="街、路段")),
    "高雄市": CountyConfig("高雄市", "64", 172400, _ZH),

    # ── 縣市 ──────────────────────────────────────────────────────────────────
    "新竹縣": CountyConfig("新竹縣", "10004", 172380, _zh(street="街或路段")),
    "苗栗縣": CountyConfig("苗栗縣", "10005", 176511, _zh(street="街、路段")),
    "彰化縣": CountyConfig("彰化縣", "10007", 170727, _zh(street="街、路段")),
    "嘉義縣": CountyConfig("嘉義縣", "10010", 172873, _zh(street="街、路段")),
    "屏東縣": CountyConfig("屏東縣", "10013", 170847,
                          dict(town="districtCode", village="village",
                               nei="neighborhood", street="streetRoadSection",
                               lane="lane", alley="alley",
                               number="houseNumber",
                               x="coordinateX", y="coordinateY")),
    "臺東縣": CountyConfig("臺東縣", "10014", 165619,
                          _zh(street="街、路段", number="號樓",
                              x="橫坐標", y="縱坐標")),
    "花蓮縣": CountyConfig("花蓮縣", "10015", 175221, _zh(street="街、路段")),
    "澎湖縣": CountyConfig("澎湖縣", "10016", 170852, _zh(street="街（路段）")),

    # 新竹市: has town code + village, but road/lane/alley/number are merged into a
    # single 地址 field — parse it with addr_col instead of the per-part columns.
    "新竹市": CountyConfig("新竹市", "10018", 157547,
                          columns=dict(town="鄉鎮市區代碼", village="村里",
                                       nei="鄰", x="橫座標", y="縱座標"),
                          addr_col="地址"),

    # 雲林縣 / 金門縣: no CSV distribution — only XLSX/XML/JSON. Read the XLSX.
    # (雲林 publishes the dataset as several XLSX parts; all are merged.)
    "雲林縣": CountyConfig("雲林縣", "10009", 166201,
                          _zh(street="街_路段"), fmt="xlsx", multi=True),
    "金門縣": CountyConfig("金門縣", "09020", 171571, _ZH, fmt="xlsx"),

    # 基隆市: not on data.gov.tw. Civil-affairs dept publishes a monthly CSV whose
    # attachment URL changes each month, so scrape the landing page for the link.
    # Format differs: district NAME (not code) and no 省市縣市代碼 / 地區 columns.
    "基隆市": CountyConfig(
        "基隆市", "10017", None,
        columns=dict(town="鄉鎮市區", village="村里", nei="鄰",
                     street="街、路段", lane="巷", alley="弄",
                     number="號", x="橫座標", y="縱座標"),
        town_scheme="name",
        page_url="https://www.klcg.gov.tw/tw/civil/2209-292163.html",
        page_re=r'/wSite/public/Attachment/\d+/[^"\'<>]+\.csv'),

    # ── 尚未支援 ──────────────────────────────────────────────────────────────
    # 未在 data.gov.tw 發布且待查自家平台: 嘉義市、南投縣、宜蘭縣、連江縣
}


# ── Networking (full TLS verification; relax only overstrict gov-server checks) ─

class _RelaxedStrictAdapter(HTTPAdapter):
    """Full CA-chain + hostname verification, with two targeted relaxations:
    1. Clears VERIFY_X509_STRICT — the Python 3.13+ flag that rejects TWCA-issued
       gov certs lacking Subject Key Identifier (e.g. data.taipei).
    2. Sets cipher security level to 1 — allows legacy weak DH params on old gov
       servers (e.g. 嘉義縣 ws-tm.cyhg.gov.tw uses DH < 1024 bit).
    This is NOT verify=False — certificate chain and hostname are still checked."""

    def init_poolmanager(self, connections, maxsize, block=False, **kwargs):
        ctx = ssl.create_default_context()
        ctx.verify_flags &= ~ssl.VERIFY_X509_STRICT
        try:
            ctx.set_ciphers("DEFAULT:@SECLEVEL=1")
        except ssl.SSLError:
            pass  # older OpenSSL without SECLEVEL support
        self.poolmanager = PoolManager(num_pools=connections, maxsize=maxsize,
                                       block=block, ssl_context=ctx, **kwargs)


def _session() -> requests.Session:
    s = requests.Session()
    s.mount("https://", _RelaxedStrictAdapter())
    s.headers.update({"User-Agent": "Mozilla/5.0 (taiwan-address-data updater)"})
    return s


# ── Helpers ────────────────────────────────────────────────────────────────────

def fullwidth_to_halfwidth(s: str) -> str:
    return "".join(chr(ord(c) - 0xFEE0) if 0xFF01 <= ord(c) <= 0xFF5E else c for c in s)


def normalize(s) -> str:
    return fullwidth_to_halfwidth((s or "").strip())


# Characters forbidden in Windows filenames (plus control chars).
_FORBIDDEN_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def safe_filename(road: str) -> str:
    """Sanitize a road name for use as a filename. Only the filename is changed;
    the road value stored inside the CSV keeps the original text.

    Empty road -> "" so no-road addresses land in the conventional {code}-.csv
    bucket. A non-empty name made of only forbidden chars -> "_"."""
    if not (road or "").strip():
        return ""
    return _FORBIDDEN_RE.sub("_", road).rstrip(" .") or "_"


def split_section(street: str) -> tuple[str, str]:
    street = (street or "").strip()
    m = _SECTION_RE.match(street)
    if m and m.group(1):
        return m.group(1), m.group(2)
    return street, ""


_NUM_RE = re.compile(r"[0-9][0-9之\-]*號.*$")   # 號 token incl 之X / trailing 樓
_ALLEY_RE = re.compile(r"[0-9]+(?:-[0-9]+)?弄$")
_LANE_RE = re.compile(r"[0-9]+(?:-[0-9]+)?巷$")


def parse_address(addr: str) -> tuple[str, str, str, str, str]:
    """Parse a merged road-level address (e.g. 三民路２巷１號) into
    (road, section, lane, alley, number). The unit chars 巷/弄/號 are kept so the
    output matches the per-column counties. Used for 新竹市 whose source merges
    everything into one 地址 field."""
    s = fullwidth_to_halfwidth((addr or "").strip())
    number = lane = alley = ""
    m = _NUM_RE.search(s)
    if m:
        number = s[m.start():]
        s = s[:m.start()]
    m = _ALLEY_RE.search(s)
    if m:
        alley = m.group(0)
        s = s[:m.start()]
    m = _LANE_RE.search(s)
    if m:
        lane = m.group(0)
        s = s[:m.start()]
    road, section = split_section(s)
    return road, section, lane, alley, number


def town_code(new_code: str, cfg: CountyConfig) -> str:
    """Convert source town code to the DGBAS 7-digit code used by roads/ files.

    Source code formats encountered:
    - 7-digit DGBAS already (most counties): returned as-is
    - 8-digit MOI / 六都 (65000010): x10 formula  →  6500100
    - 8-digit county MOI (10007010): truncate to 7 → 1000701
    """
    code = (new_code or "").strip()
    if len(code) == 6:
        # XLSX numeric cells can drop a leading zero (金門 0902001 -> 902001)
        code = code.zfill(7)
    if len(code) == 7:
        return code  # already DGBAS
    if len(code) == 8:
        if cfg.town_scheme == "x10" or len(cfg.code) == 2:
            # 六都: multiply seq by 10 to get DGBAS last 4 digits
            seq = int(code[-3:])
            return cfg.code + "0" + str(seq * 10).zfill(4)
        # County with 5-digit code (e.g. 10007): truncate last char
        return code[:7]
    raise ValueError(f"unexpected town code length={len(code)}: {code!r}")


def load_area(cfg: CountyConfig) -> tuple[dict, dict]:
    """From area_2014.csv build:
      town_names: old_town_code -> district display name (e.g. 板橋區)
      villages  : (old_town_code, village_name) -> village_code
    Fully derived from data, no hardcoded district tables."""
    prefix = cfg.name
    town_names: dict[str, str] = {}
    villages: dict[tuple[str, str], str] = {}
    with open(AREA_CSV, encoding="utf-8") as f:
        for row in csv.reader(f):
            if len(row) < 2:
                continue
            # Some single-char districts are padded with a full-width space to
            # align width (東區 -> 東　區). Strip it so names are consistent.
            name, code = row[0].strip().replace("　", ""), row[1].strip()
            if not name.startswith(prefix):
                continue
            rest = name[len(prefix):]
            if "-" not in code:
                # district row: "新北市板橋區" -> 6500100
                if rest and len(code) == 7:
                    town_names[code] = rest
            else:
                # village row: "新北市板橋區港嘴里" -> 6500100-050
                tcode = code.split("-")[0]
                district = town_names.get(tcode, "")
                if district and rest.startswith(district):
                    village = rest[len(district):]
                    if village:
                        villages[(tcode, village)] = code
    return town_names, villages


# ── Dataset resolution / download ──────────────────────────────────────────────

def get_download_urls(cfg: CountyConfig, s) -> list[str]:
    """Resolve all download URLs matching cfg.fmt from the data.gov.tw API.
    Returns a list (a dataset may publish several parts, e.g. 雲林 XLSX)."""
    want = cfg.fmt.upper()
    domains = [cfg.domain] + [d for d in ("data.gov.tw", "data.nat.gov.tw")
                               if d != cfg.domain]
    last = None
    for dom in domains:
        try:
            r = s.get(f"https://{dom}/api/v2/rest/dataset/{cfg.dataset_id}",
                      timeout=30)
            if r.status_code == 200 and r.text.strip().startswith("{"):
                res = r.json().get("result")
                if res:
                    urls, seen = [], set()
                    for d in res.get("distribution", []):
                        if (d.get("resourceFormat", "").upper() == want
                                and d.get("resourceDownloadUrl")):
                            u = d["resourceDownloadUrl"].strip()
                            if u not in seen:
                                seen.add(u)
                                urls.append(u)
                    if urls:
                        return urls
        except Exception as e:
            last = e
    raise RuntimeError(f"no {want} resource for dataset {cfg.dataset_id} ({last})")


def resolve_sources(cfg: CountyConfig, s) -> list[str]:
    """Return the download URL(s) for a county, honoring url / page_url / API."""
    if cfg.url:
        return [cfg.url]
    if cfg.page_url:
        from urllib.parse import urljoin
        r = s.get(cfg.page_url, timeout=30)
        r.raise_for_status()
        hits = re.findall(cfg.page_re, r.text)
        if not hits:
            raise RuntimeError(f"no link matching {cfg.page_re!r} on {cfg.page_url}")
        return [urljoin(cfg.page_url, hits[0])]
    return get_download_urls(cfg, s)


def download_bytes(url: str, s) -> bytes:
    print(f"Downloading:\n  {url}")
    r = s.get(url, timeout=600, stream=True)
    r.raise_for_status()
    chunks, total = [], 0
    for chunk in r.iter_content(chunk_size=1024 * 1024):
        chunks.append(chunk)
        total += len(chunk)
        print(f"\r  {total/1024/1024:.1f} MB", end="", flush=True)
    print()
    return b"".join(chunks)


def decode_bytes(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "big5hkscs", "big5", "cp950"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("big5hkscs", errors="replace")


def read_csv_rows(raw: bytes) -> list[dict]:
    reader = csv.DictReader(io.StringIO(decode_bytes(raw)))
    if reader.fieldnames:  # some counties pad header names with spaces
        reader.fieldnames = [f.strip() for f in reader.fieldnames]
    return list(reader)


def read_xlsx_rows(raw: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it)]
    rows = []
    for r in it:
        d = {}
        for i, v in enumerate(r):
            if i < len(header) and header[i]:
                if v is None:
                    v = ""
                elif isinstance(v, float) and v.is_integer():
                    v = str(int(v))   # avoid "1000911.0" for code/number cells
                else:
                    v = str(v)
                d[header[i]] = v
        rows.append(d)
    wb.close()
    return rows


def load_rows(cfg: CountyConfig, s) -> list[dict]:
    urls = resolve_sources(cfg, s)
    if not cfg.multi and len(urls) > 1:
        # dataset lists several redundant copies (e.g. 臺南) — use only the first
        urls = urls[:1]
    rows: list[dict] = []
    for url in urls:
        raw = download_bytes(url, s)
        rows.extend(read_xlsx_rows(raw) if cfg.fmt == "xlsx"
                    else read_csv_rows(raw))
    if len(urls) > 1:
        # several parts merged — drop any exactly-duplicate rows across files
        seen, uniq = set(), []
        for r in rows:
            k = tuple(sorted(r.items()))
            if k not in seen:
                seen.add(k)
                uniq.append(r)
        print(f"  merged {len(urls)} files: {len(rows):,} -> {len(uniq):,} rows")
        rows = uniq
    return rows


# ── Pipeline ───────────────────────────────────────────────────────────────────

def run(cfg: CountyConfig) -> None:
    try:
        from pyproj import Transformer
    except ImportError:
        sys.exit("ERROR: pyproj required.  pip install pyproj")

    transformer = Transformer.from_crs(cfg.epsg, "EPSG:4326", always_xy=True)
    s = _session()

    src = cfg.dataset_id if cfg.dataset_id else "direct"
    print(f"=== Updating {cfg.name} (county {cfg.code}, source {src}) ===")

    print("Loading area lookup…")
    town_names, villages = load_area(cfg)
    print(f"  {len(town_names)} districts, {len(villages)} villages")
    # district name -> DGBAS code, for counties that give the name not the code
    name_to_code = {v: k for k, v in town_names.items()}

    rows = load_rows(cfg, s)

    col = cfg.columns
    road_groups: dict[str, list[list]] = defaultdict(list)
    skipped = bad_coords = unmatched_villages = 0

    print("Processing records…")
    for i, row in enumerate(rows, 1):
        try:
            if cfg.town_scheme == "name":
                tcode = name_to_code.get((row.get(col["town"]) or "").strip())
                if not tcode:
                    raise ValueError("unknown district name")
            else:
                tcode = town_code((row.get(col["town"]) or "").strip(), cfg)
            village_name = (row.get(col["village"]) or "").strip()
            village = villages.get((tcode, village_name), tcode + "-000")
            if village == tcode + "-000":
                unmatched_villages += 1

            nei_raw = (row.get(col["nei"]) or "").strip().lstrip("0") or "0"
            neighborhood = (nei_raw + "鄰") if nei_raw != "0" else ""

            if cfg.addr_col:
                road, section, lane, alley, number = parse_address(
                    row.get(cfg.addr_col))
            else:
                road, section = split_section((row.get(col["street"]) or "").strip())
                lane = normalize(row.get(col["lane"]))
                alley = normalize(row.get(col["alley"]))
                number = normalize(row.get(col["number"]))

            x_raw = (row.get(col["x"]) or "").strip()
            y_raw = (row.get(col["y"]) or "").strip()
            lon, lat = transformer.transform(float(x_raw), float(y_raw))
        except (KeyError, ValueError, TypeError):
            skipped += 1
            continue

        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
            bad_coords += 1
            continue

        district = town_names.get(tcode, "")
        full_addr = (cfg.name + district + village_name + neighborhood
                     + road + section + lane + alley + number)
        record = [full_addr, cfg.code, tcode, village, neighborhood,
                  road, section, lane, alley, "", "", number,
                  f"{lon:.15g}", f"{lat:.15g}"]
        road_groups[road or ""].append(record)
        if i % 100_000 == 0:
            print(f"  {i:,} processed…")

    total = sum(len(v) for v in road_groups.values())
    print(f"  Done — {total:,} records, {skipped} skipped, "
          f"{bad_coords} bad coords, {unmatched_villages} unmatched villages "
          f"({unmatched_villages/max(total,1)*100:.2f}%)")
    if total == 0:
        sys.exit("ERROR: 0 records produced — aborting without touching roads/ "
                 "(check column mapping for this county).")

    print(f"Removing old {cfg.code}-*.csv…")
    removed = sum(1 for f in ROADS_DIR.glob(f"{cfg.code}-*.csv")
                  if (f.unlink() or True))
    print(f"  Removed {removed} files")

    # Merge by sanitized filename so Windows-illegal road names (e.g. containing
    # ? * : ) don't crash, and any two roads colliding to one name are combined.
    by_file: dict[str, list[list]] = defaultdict(list)
    collisions = 0
    for road, records in road_groups.items():
        fname = safe_filename(road)
        if fname != road and by_file[fname]:
            collisions += 1
        by_file[fname].extend(records)

    print(f"Writing {len(by_file)} road files…")
    for fname, records in by_file.items():
        with open(ROADS_DIR / f"{cfg.code}-{fname}.csv", "w",
                  newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(CSV_HEADER)
            w.writerows(records)
    note = f" ({collisions} sanitized-name merges)" if collisions else ""
    print(f"Done. {cfg.name}: {len(by_file)} road files, {total:,} records{note}.")


def main():
    ap = argparse.ArgumentParser(description="Update Taiwan county door-plate map data")
    ap.add_argument("county", nargs="?", help="county name, e.g. 新北市")
    ap.add_argument("--list", action="store_true", help="list configured counties")
    ap.add_argument("--all", action="store_true", help="run all configured counties")
    args = ap.parse_args()

    if args.list or (not args.county and not args.all):
        print("Configured counties:")
        for name, c in COUNTIES.items():
            src = f"dataset {c.dataset_id}" if c.dataset_id else "direct CSV"
            tag = f", {c.fmt}" if c.fmt != "csv" else ""
            print(f"  {name}  (code {c.code}, {src}{tag})")
        return

    if args.all:
        for name in COUNTIES:
            run(COUNTIES[name])
        return

    if args.county not in COUNTIES:
        sys.exit(f"Unknown county {args.county!r}. Use --list to see configured ones.")
    run(COUNTIES[args.county])


if __name__ == "__main__":
    main()
